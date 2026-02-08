#!/usr/bin/env python3
from __future__ import annotations

import importlib
import json
import sys
import traceback
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import List

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from _scripts.fig_renderers.common import FIG_XLSX, EXPORTS_DIR, RenderContext
from _scripts.fig_renderers.registry import FIGURE_RENDERER_MODULES


@dataclass
class RenderResult:
    module: str
    output_file: str
    tab: str
    ok: bool
    error: str = ""


def main() -> int:
    if not FIG_XLSX.exists():
        print(f"ERROR: Missing source-of-truth workbook: {FIG_XLSX}")
        return 2

    wb = load_workbook(FIG_XLSX, data_only=True)
    ctx = RenderContext(wb=wb)

    results: List[RenderResult] = []

    print(f"Building figures from workbook: {FIG_XLSX}")
    print(f"Export target directory: {EXPORTS_DIR}")
    print(f"Renderers configured: {len(FIGURE_RENDERER_MODULES)}")

    for mod_name in FIGURE_RENDERER_MODULES:
        try:
            mod = importlib.import_module(mod_name)
        except Exception as exc:
            results.append(RenderResult(module=mod_name, output_file="(unknown)", tab="(unknown)", ok=False, error=f"import failed: {exc}"))
            break

        spec = getattr(mod, "SPEC", None)
        if not isinstance(spec, dict):
            results.append(RenderResult(module=mod_name, output_file="(unknown)", tab="(unknown)", ok=False, error="missing SPEC dict"))
            break

        tab = str(spec.get("tab", ""))
        out = str(spec.get("output_file", ""))

        if not tab or not out:
            results.append(RenderResult(module=mod_name, output_file=out or "(missing)", tab=tab or "(missing)", ok=False, error="SPEC requires tab and output_file"))
            break

        if tab not in wb.sheetnames:
            results.append(RenderResult(module=mod_name, output_file=out, tab=tab, ok=False, error=f"missing sheet '{tab}'"))
            break

        try:
            rendered_path = mod.render(ctx)
            rendered_path = Path(rendered_path)
            if not rendered_path.exists() or rendered_path.stat().st_size == 0:
                raise RuntimeError(f"renderer returned missing/empty file: {rendered_path}")
            results.append(RenderResult(module=mod_name, output_file=out, tab=tab, ok=True))
            print(f"  OK  {out} <- {tab} ({mod_name.split('.')[-1]})")
        except Exception as exc:
            tb = traceback.format_exc(limit=4)
            results.append(RenderResult(module=mod_name, output_file=out, tab=tab, ok=False, error=f"{exc}\n{tb}"))
            break

    ok_count = sum(1 for r in results if r.ok)
    fail = next((r for r in results if not r.ok), None)

    print("\nSummary")
    print(f"  Rendered OK: {ok_count}")
    print(f"  Failed: {0 if fail is None else 1}")

    qa_dir = ROOT / "_output" / "qa"
    qa_dir.mkdir(parents=True, exist_ok=True)
    out_json = qa_dir / "figure_render_last_run.json"
    out_json.write_text(json.dumps([asdict(r) for r in results], indent=2), encoding="utf-8")
    print(f"  Log: {out_json}")

    if fail:
        print("\nFAIL-FAST ERROR")
        print(f"  Module: {fail.module}")
        print(f"  Output: {fail.output_file}")
        print(f"  Tab: {fail.tab}")
        print(f"  Error: {fail.error}")
        return 1

    # strict post-check: every output file listed in modules must exist
    missing = []
    for mod_name in FIGURE_RENDERER_MODULES:
        mod = importlib.import_module(mod_name)
        out = mod.SPEC["output_file"]
        p = EXPORTS_DIR / out
        if not p.exists() or p.stat().st_size == 0:
            missing.append(str(p))

    if missing:
        print("\nERROR: Strict post-check failed. Missing outputs:")
        for m in missing:
            print(f"  - {m}")
        return 1

    print("\nAll configured figures regenerated successfully.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
