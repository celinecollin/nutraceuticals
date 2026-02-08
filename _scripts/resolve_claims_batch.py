import pandas as pd
from openpyxl import load_workbook
import os
import re

REGISTRY_PATH = '_registry/source_registry.xlsx'
SECTIONS_DIR = 'sections'

NEW_SOURCES = [
    {'id': 'S122', 'filename': 'sources/articles/FeedAdditive_ROI_3to1.txt', 'origin': 'Feed & Additive Magazine', 'title': 'Phytogenic Feed Additives ROI', 'date': '2026-02-08'},
    {'id': 'S123', 'filename': 'sources/articles/PetFoodInd_UrbanSuburban.txt', 'origin': 'Petfood Industry', 'title': 'Urban vs Suburban Purchasing Habits', 'date': '2026-02-08'},
    {'id': 'S124', 'filename': 'sources/regulatory/MARA_Announcement_194_Summary.txt', 'origin': 'MARA China', 'title': 'China AGP Ban Announcement 194', 'date': '2020-07-01'},
    {'id': 'S125', 'filename': 'sources/reports/Sector_Deal_Multiples_2020-2024.txt', 'origin': 'Public Financial Data', 'title': 'Sector Deal Multiples Assessment', 'date': '2026-02-08'},
    {'id': 'S126', 'filename': 'sources/regulatory/EU_Green_Claims_Directive_Summary.txt', 'origin': 'EU Commission', 'title': 'Green Claims Directive Proposal', 'date': '2023-03-22'},
    {'id': 'S127', 'filename': 'sources/academic/Nutrigenomics_Review_Summary.txt', 'origin': 'Frontiers / NIH', 'title': 'Nutrigenomics Review', 'date': '2026-02-08'}
]

CLAIM_UPDATES = {
    'C063': 'S122',
    'C061': 'S123',
    'C076': 'S124',
    'C072': 'S125',
    'C073': 'S125', # Valuation dispersion supported by multiples range
    'C084': 'S125', # Validates Blue Buffalo, Erber, etc.
    'C086': 'S125', # Two-speed multiple
    'C074': 'S126',
    'C075': 'S127'
}

def update_registry():
    print("Updating registry...")
    # Load workbook using openpyxl to append rows safely
    wb = load_workbook(REGISTRY_PATH)
    
    # Update Sources
    ws_sources = wb['Sources']
    # Find next empty row
    row = ws_sources.max_row + 1
    for src in NEW_SOURCES:
        ws_sources.cell(row=row, column=1, value=src['id'])
        ws_sources.cell(row=row, column=2, value=src['filename'])
        ws_sources.cell(row=row, column=3, value=src['origin'])
        ws_sources.cell(row=row, column=4, value=src['title'])
        ws_sources.cell(row=row, column=5, value=src['date'])
        row += 1
    
    # Update Claims
    ws_claims = wb['Claims']
    # Iterate rows to find Claim IDs
    header = [cell.value for cell in ws_claims[1]]
    id_col_idx = header.index('claim_id') + 1
    src_col_idx = header.index('source_ids') + 1
    status_col_idx = header.index('verified') + 1
    
    count = 0
    for r in range(2, ws_claims.max_row + 1):
        c_id = ws_claims.cell(row=r, column=id_col_idx).value
        if c_id in CLAIM_UPDATES:
            new_sid = CLAIM_UPDATES[c_id]
            current_src = ws_claims.cell(row=r, column=src_col_idx).value
            
            # Replace [UNVERIFIED] with S-tag, or append if other sources exist
            if 'UNVERIFIED' in str(current_src):
                new_val = str(current_src).replace('UNVERIFIED', new_sid)
                if new_val.startswith(', '): new_val = new_val[2:]
                ws_claims.cell(row=r, column=src_col_idx, value=new_val)
                ws_claims.cell(row=r, column=status_col_idx, value='Y') # Mark as verified? Or keep N? User said "keep verified=N" in request!
                # Wait, request said: "Claims tab: update source_ids/source_location/notes, keep verified=N"
                # So I will NOT set verified=Y.
                ws_claims.cell(row=r, column=status_col_idx, value='N') 
                count += 1
    
    wb.save(REGISTRY_PATH)
    print(f"Registry updated: Added {len(NEW_SOURCES)} sources, updated {count} claims.")

def patch_sections():
    print("Patching sections...")
    # Map C-ID to regex for [UNVERIFIED] replacement
    # We need to read sections and find where these claims are. 
    # Since we don't have line numbers, we'll replace [UNVERIFIED] based on context or just replace specific [UNVERIFIED] blocks if we can identify them?
    # Better approach: Read UNVERIFIED_CLAIMS.md to see which *section* files they are in, but the prompt says "Files to update: sections/*.md".
    # I will iterate all section files and grep for the claim context I extracted earlier.
    # Actually, I have the claim text from Step 18. I can search for unique substrings of the claim text.
    
    # Simplified mapping based on unique text snippets from Step 18
    snippets = {
        'C063': {'text': '3:1 economic hurdle', 'tag': '[S122]'},
        'C061': {'text': 'channel-demography coupling', 'tag': '[S123]'},
        'C076': {'text': 'APAC AGP-dividend', 'tag': '[S124]'},
        'C072': {'text': 'transaction evidence supports this consolidation', 'tag': '[S125]'},
        'C073': {'text': 'Valuation dispersion remains linked', 'tag': '[S125]'},
        'C084': {'text': 'Legacy v19 transaction history', 'tag': '[S125]'},
        'C086': {'text': 'two-speed multiple framing', 'tag': '[S125]'},
        'C074': {'text': 'Green-claim institutionalization', 'tag': '[S126]'},
        'C075': {'text': 'longevity-linked protocols', 'tag': '[S127]'}
    }
    
    files = [f for f in os.listdir(SECTIONS_DIR) if f.endswith('.md')]
    
    for f in files:
        path = os.path.join(SECTIONS_DIR, f)
        with open(path, 'r') as file:
            content = file.read()
        
        modified = False
        for c_id, data in snippets.items():
            # We look for the text. If found, we look for [UNVERIFIED] nearby and replace it.
            # This is tricky with regex. 
            # Alternative: Search for the text, then replace the NEXT occurrence of [UNVERIFIED].
            if data['text'] in content:
                # Find the location
                idx = content.find(data['text'])
                # Look ahead for [UNVERIFIED]
                # We'll replace the first [UNVERIFIED] that appears after the snippet
                # But we must be careful not to replace it if it's too far (next paragraph).
                # Let's assume it's in the same paragraph.
                
                # Split content into parts
                pre = content[:idx]
                post = content[idx:]
                
                if '[UNVERIFIED]' in post:
                    # Check distance
                    u_idx = post.find('[UNVERIFIED]')
                    if u_idx < 1000: # Arbitrary "paragraph" length
                        # Replace
                        new_post = post.replace('[UNVERIFIED]', data['tag'], 1)
                        content = pre + new_post
                        modified = True
                        print(f"Patched {c_id} in {f}")
        
        if modified:
            with open(path, 'w') as file:
                file.write(content)

if __name__ == "__main__":
    update_registry()
    patch_sections()
