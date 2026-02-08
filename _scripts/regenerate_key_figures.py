#!/usr/bin/env python3
from __future__ import annotations

import math
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
EXPORTS = ROOT / "_figures" / "exports"
FIG_XLSX = ROOT / "_figures" / "figures_data.xlsx"

plt.rcParams["font.family"] = "DejaVu Sans"


def save(fig, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.tight_layout()
    fig.savefig(path, dpi=220, bbox_inches="tight")
    plt.close(fig)


def regen_figure_1() -> None:
    rows = [
        ["Feature", "United States (US)", "European Union (EU)", "United Kingdom (UK)", "China"],
        ["Nutraceutical Definition", "Undefined (Food vs Drug)", "Undefined (Feed vs VMP)", "Undefined (Feed vs VMP)", "Undefined (Feed Additive vs VMP)"],
        ["Regulatory Body", "FDA-CVM & AAFCO", "EFSA & National Agencies", "VMD & FSA", "MARA & SAMR"],
        ["Disease Claims", "Prohibited (Drug only)", "Prohibited (PARNUTs exception)", "Prohibited (Drug only)", "Restricted by registration class"],
        ["Market Entry Speed", "Fast (Notification)", "Slow (Dossier Approval)", "Medium (Route-dependent)", "Medium-Slow (Registration dependent)"],
    ]

    df = pd.DataFrame(rows[1:], columns=rows[0])
    for out in [ROOT / "sources" / "datasets" / "Table_US_vs_EU.csv", EXPORTS / "Table_US_vs_EU.csv"]:
        df.to_csv(out, index=False)

    fig, ax = plt.subplots(figsize=(16, 8))
    ax.axis("off")
    ax.set_title("Regulatory Landscape Comparison: US vs EU vs UK vs China", fontsize=22, fontweight="bold", pad=18)

    table = ax.table(cellText=rows[1:], colLabels=rows[0], cellLoc="left", colLoc="left", loc="center")
    table.auto_set_font_size(False)
    table.set_fontsize(11)
    table.scale(1.1, 2.0)

    for (r, c), cell in table.get_celld().items():
        cell.set_edgecolor("#d9d9d9")
        if r == 0:
            cell.set_facecolor("#0b3a63")
            cell.set_text_props(color="white", fontweight="bold")
        else:
            cell.set_facecolor("#f7f9fc" if r % 2 == 1 else "#ffffff")
            if c == 0:
                cell.set_text_props(fontweight="bold")

    save(fig, EXPORTS / "Table_US_vs_EU.png")


def _read_fig5_df() -> pd.DataFrame:
    wb = load_workbook(FIG_XLSX, data_only=True)
    ws = wb["Figure 5"]
    rows = list(ws.iter_rows(min_row=3, max_col=4, values_only=True))
    clean = [r for r in rows if r[0] and isinstance(r[1], (int, float)) and isinstance(r[2], (int, float))]
    df = pd.DataFrame(clean, columns=["Company", "R&D Intensity (%)", "EBITDA Margin (%)", "Sector"])
    return df


def regen_figure_5() -> None:
    df = _read_fig5_df()

    colors = {
        "Pharma-Nutra": "#0b3a63",
        "Ingredient Tech": "#1f77b4",
        "Brand/Consumer": "#2ca02c",
        "Feed/Health": "#ff7f0e",
        "Commodity Feed": "#d62728",
    }

    fig, ax = plt.subplots(figsize=(13.5, 8.8))
    for sector, part in df.groupby("Sector"):
        ax.scatter(
            part["R&D Intensity (%)"],
            part["EBITDA Margin (%)"],
            s=90,
            alpha=0.9,
            label=sector,
            color=colors.get(sector, "#7f7f7f"),
            edgecolor="white",
            linewidth=0.8,
        )

    for _, row in df.iterrows():
        ax.annotate(
            row["Company"],
            (row["R&D Intensity (%)"], row["EBITDA Margin (%)"]),
            textcoords="offset points",
            xytext=(5, 4),
            fontsize=8.6,
        )

    ax.axvline(5, color="#4c4c4c", linestyle="--", linewidth=1)
    ax.axhline(20, color="#4c4c4c", linestyle="--", linewidth=1)

    ax.text(9.6, 39, "High R&D / High Margin", fontsize=9, color="#0b3a63", fontweight="bold")
    ax.text(1.0, 4.5, "Low R&D / Low Margin", fontsize=9, color="#6b6b6b")

    ax.set_title("The Innovation-Premium Matrix: R&D Intensity vs EBITDA Margin", fontsize=18, fontweight="bold", pad=12)
    ax.set_xlabel("R&D Intensity (% of Revenue)", fontsize=11)
    ax.set_ylabel("EBITDA Margin (%)", fontsize=11)
    ax.set_xlim(0, max(12, math.ceil(df["R&D Intensity (%)"].max()) + 1))
    ax.set_ylim(0, max(42, math.ceil(df["EBITDA Margin (%)"].max()) + 2))
    ax.grid(axis="both", linestyle=":", alpha=0.35)
    ax.legend(title="Sector", fontsize=8.5, title_fontsize=9.5, frameon=True, loc="lower right")

    save(fig, EXPORTS / "Figure_II_0_1_Innovation_Matrix.png")


def regen_figure_16() -> None:
    data = pd.DataFrame(
        {
            "Region": ["US", "Canada", "EU"],
            "Ownership Rate (%)": [71, 60, 49],
        }
    )
    for out in [ROOT / "sources" / "datasets" / "Figure1_Pet_Ownership.csv", EXPORTS / "Figure1_Pet_Ownership.csv"]:
        data.to_csv(out, index=False)

    fig, ax = plt.subplots(figsize=(11.5, 6.3))
    bars = ax.bar(data["Region"], data["Ownership Rate (%)"], color=["#0b3a63", "#2e6ea3", "#7aa6d1"])
    for b in bars:
        h = b.get_height()
        ax.text(b.get_x() + b.get_width() / 2, h + 1.1, f"{h:.0f}%", ha="center", va="bottom", fontsize=10)

    ax.set_title("Pet Ownership Rates in Developed Markets", fontsize=18, fontweight="bold")
    ax.set_xlabel("Region")
    ax.set_ylabel("Pet Ownership Rate (%)")
    ax.set_ylim(0, 80)
    ax.grid(axis="y", linestyle=":", alpha=0.35)

    save(fig, EXPORTS / "Figure1_Pet_Ownership.png")


def regen_figure_35() -> None:
    csv_path = EXPORTS / "Figure21_Pharma_Funnel.csv"
    df = pd.read_csv(csv_path)
    stage_col = df.columns[0]
    value_col = df.columns[1]

    fig, ax = plt.subplots(figsize=(10.5, 6.2))
    y = np.arange(len(df))
    ax.barh(y, df[value_col], color="#0b3a63")
    ax.set_yticks(y, labels=df[stage_col])
    ax.invert_yaxis()
    ax.set_title("Pharma Integration Funnel", fontsize=18, fontweight="bold")
    ax.set_xlabel("Relative Integration Potential (Index, Trial = 100)")
    ax.set_ylabel("Stage")
    ax.grid(axis="x", linestyle=":", alpha=0.35)

    for i, v in enumerate(df[value_col]):
        ax.text(v + 1.2, i, f"{v:.0f}", va="center", fontsize=10)

    save(fig, EXPORTS / "Figure21_Pharma_Funnel.png")


def regen_simplified_ii_1() -> None:
    species = ["Dogs", "Cats", "Poultry", "Swine", "Ruminants", "Aquaculture"]
    themes = ["Mobility", "Behavior", "Gut/Immunity", "Performance/FCR", "Sustainability"]
    matrix = np.array(
        [
            [9, 7, 6, 2, 3],
            [7, 8, 6, 1, 2],
            [1, 1, 8, 9, 7],
            [1, 1, 8, 8, 7],
            [2, 1, 7, 8, 9],
            [1, 1, 7, 8, 8],
        ],
        dtype=float,
    )

    fig, ax = plt.subplots(figsize=(11.8, 6.5))
    im = ax.imshow(matrix, cmap="Blues", vmin=0, vmax=10)
    ax.set_xticks(np.arange(len(themes)), labels=themes)
    ax.set_yticks(np.arange(len(species)), labels=species)
    ax.set_title("Figure II.1 (Simplified): Functional Needs by Species", fontsize=16, fontweight="bold")
    ax.set_xlabel("Functional Theme")
    ax.set_ylabel("Primary Species Group")

    for i in range(matrix.shape[0]):
        for j in range(matrix.shape[1]):
            val = int(matrix[i, j])
            ax.text(j, i, str(val), ha="center", va="center", color=("white" if val >= 7 else "#0b3a63"), fontsize=9)

    cbar = fig.colorbar(im, ax=ax, fraction=0.035, pad=0.02)
    cbar.set_label("Relative Strategic Relevance (0-10)")

    save(fig, EXPORTS / "Figure_II_1_Simplified.png")


def regen_simplified_ii_12() -> None:
    segments = ["Mobility", "Cognition", "Behavior", "Gut Health", "Immunity", "Performance/FCR"]
    totals = np.array([776, 312, 239, 2913, 1841, 1426], dtype=float)
    pet_share = np.array([0.92, 0.95, 0.96, 0.45, 0.55, 0.10])
    livestock_share = 1 - pet_share
    pet_vals = totals * pet_share
    livestock_vals = totals * livestock_share

    x = np.arange(len(segments))
    width = 0.62

    fig, ax = plt.subplots(figsize=(12.2, 6.8))
    ax.bar(x, pet_vals, width=width, label="Companion/Pet-led", color="#1f77b4")
    ax.bar(x, livestock_vals, width=width, bottom=pet_vals, label="Production/Livestock-led", color="#ff7f0e")

    ax.set_title("Figure II.12 (Simplified): Comparative Economic Value by Segment", fontsize=16, fontweight="bold")
    ax.set_xlabel("Functional Segment")
    ax.set_ylabel("Estimated Value (USD millions)")
    ax.set_xticks(x, labels=segments, rotation=15)
    ax.grid(axis="y", linestyle=":", alpha=0.35)
    ax.legend(loc="upper right")

    for i, t in enumerate(totals):
        ax.text(i, t + 35, f"{int(t):,}", ha="center", va="bottom", fontsize=9)

    save(fig, EXPORTS / "Figure_II_12_Simplified.png")


def main() -> None:
    regen_figure_1()
    regen_figure_5()
    regen_figure_16()
    regen_figure_35()
    regen_simplified_ii_1()
    regen_simplified_ii_12()
    print("Regenerated: Figure 1, Figure 5, Figure 16, Figure 35, Figure II.1 (simplified), Figure II.12 (simplified)")


if __name__ == "__main__":
    main()
