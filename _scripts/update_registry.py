

from openpyxl import load_workbook
import sys
import json
import os
from datetime import datetime

def append_claims(registry_path, claims_data):
    """Appends new claims to the Claims tab of the registry."""
    try:
        # Load the workbook
        book = load_workbook(registry_path)
        if 'Claims' not in book.sheetnames:
            return {"error": "Claims tab not found"}
        
        sheet = book['Claims']
        
        # Find next available row
        next_row = sheet.max_row + 1
        
        # Get existing claim IDs to generate new ones
        existing_ids = []
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
            if row[0]:
                existing_ids.append(row[0])
        
        last_id_num = 0
        if existing_ids:
            try:
                # expecting C001, C002, etc.
                last_id_num = max([int(cid[1:]) for cid in existing_ids if cid.startswith('C') and cid[1:].isdigit()])
            except ValueError:
                pass

        new_claims = []
        for i, claim in enumerate(claims_data):
            claim_id = f"C{last_id_num + 1 + i:03d}"
            
            # Map dictionary keys to columns (based on PRD)
            # A: claim_id, B: section, C: claim_text, D: source_ids, E: source_location, 
            # F: verified, G: agent_generated, H: date_added, I: date_verified, J: notes
            
            row_data = [
                claim_id,
                claim.get('section', ''),
                claim.get('claim_text', ''),
                claim.get('source_ids', ''),
                claim.get('source_location', ''),
                claim.get('verified', 'N'),
                claim.get('agent_generated', 'Y'),
                datetime.now().strftime('%Y-%m-%d'),
                '', # date_verified
                claim.get('notes', '')
            ]
            
            # Append to sheet
            sheet.append(row_data)
            new_claims.append(claim_id)

        book.save(registry_path)
        return {"success": True, "added_claims": new_claims}

    except Exception as e:
        return {"error": str(e)}

if __name__ == "__main__":
    registry_path = "_registry/source_registry.xlsx"
    if len(sys.argv) > 1:
        # If the first arg is not a json string, treat it as path
        if not sys.argv[1].strip().startswith('['):
             registry_path = sys.argv[1]
             json_input = sys.argv[2] if len(sys.argv) > 2 else "[]"
        else:
             json_input = sys.argv[1]
    else:
        # Read from stdin if no args
        json_input = sys.stdin.read()

    try:
        claims_data = json.loads(json_input)
    except json.JSONDecodeError:
        print(json.dumps({"error": "Invalid JSON input"}))
        sys.exit(1)
        
    if not os.path.exists(registry_path):
        print(json.dumps({"error": f"File not found: {registry_path}"}))
        sys.exit(1)

    result = append_claims(registry_path, claims_data)
    print(json.dumps(result, indent=2))
