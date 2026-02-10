from __future__ import annotations

from dataclasses import dataclass
import io
import json
from pathlib import Path
from typing import Any, Dict, List, Optional

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(__file__).resolve().parents[2]
FIG_XLSX = ROOT / "_figures" / "figures_data.xlsx"
EXPORTS_DIR = ROOT / "_figures" / "exports"
LEGACY_DIMENSIONS_JSON = ROOT / "_scripts" / "fig_renderers" / "legacy_dimensions.json"


def _load_legacy_dimensions() -> Dict[str, tuple[int, int]]:
    if not LEGACY_DIMENSIONS_JSON.exists():
        return {}
    try:
        raw = json.loads(LEGACY_DIMENSIONS_JSON.read_text(encoding="utf-8"))
    except Exception:
        return {}
    dims: Dict[str, tuple[int, int]] = {}
    for name, meta in raw.items():
        if not isinstance(meta, dict):
            continue
        try:
            dims[str(name)] = (int(meta["width"]), int(meta["height"]))
        except Exception:
            continue
    return dims


def _load_dimensions_from_baseline_commit() -> Dict[str, tuple[int, int]]:
    # Fast, local fallback map captured during refactor to preserve legacy insert sizing in DOCX.
    path = ROOT / "_output" / "qa" / "figures_after_refactor_comparison.json"
    if not path.exists():
        return {}
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}

    dims: Dict[str, tuple[int, int]] = {}
    for row in raw.get("results", []):
        if not isinstance(row, dict):
            continue
        name = row.get("file")
        baseline_dims = row.get("baseline_dims")
        if not name or not isinstance(baseline_dims, list) or len(baseline_dims) != 2:
            continue
        try:
            dims[str(name)] = (int(baseline_dims[0]), int(baseline_dims[1]))
        except Exception:
            continue
    return dims


LEGACY_DIMS = _load_legacy_dimensions() or _load_dimensions_from_baseline_commit()


def setup_style() -> None:
    plt.rcParams["font.family"] = "DejaVu Sans"
    plt.rcParams["axes.titlesize"] = 18
    plt.rcParams["axes.labelsize"] = 11
    plt.rcParams["xtick.labelsize"] = 9
    plt.rcParams["ytick.labelsize"] = 9


def save_figure(fig: plt.Figure, output_file: str, dpi: int = 220) -> Path:
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    out = EXPORTS_DIR / output_file
    fig.tight_layout()
    fig.savefig(out, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    _enforce_legacy_dimensions(out, output_file)
    return out


def _enforce_legacy_dimensions(out: Path, output_file: str) -> None:
    # Keep exact legacy dimensions where known so downstream DOCX layout remains stable.
    target = LEGACY_DIMS.get(output_file)
    if not target:
        return
    try:
        with Image.open(out) as im:
            if im.size != target:
                resampling = getattr(Image, "Resampling", Image).LANCZOS
                im = im.resize(target, resampling)
                im.save(out)
    except Exception:
        # Do not mask the main render pass if resize fails on a single output.
        pass


def _find_header_row(ws) -> int:
    # Most tabs use row 3; Figure 1 uses row 2.
    for r in range(2, 8):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 16)]
        non_empty = [v for v in vals if v not in (None, "")]
        if len(non_empty) >= 2 and not str(non_empty[0]).lower().startswith("source:"):
            return r
    return 3


def read_df_from_tab(wb: Workbook, tab_name: str) -> pd.DataFrame:
    if tab_name not in wb.sheetnames:
        raise ValueError(f"Missing tab '{tab_name}' in figures_data.xlsx")

    ws = wb[tab_name]
    header_row = _find_header_row(ws)

    headers: List[str] = []
    col = 1
    while col <= 30:
        v = ws.cell(row=header_row, column=col).value
        if v in (None, ""):
            if headers:
                break
            col += 1
            continue
        headers.append(str(v).strip())
        col += 1

    if not headers:
        raise ValueError(f"No headers found in tab '{tab_name}'")

    rows: List[List[Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        first = vals[0]

        if first in (None, "") and all(v in (None, "") for v in vals[1:]):
            break

        if isinstance(first, str) and first.strip().lower().startswith("source:"):
            break

        rows.append(vals)

    df = pd.DataFrame(rows, columns=headers)
    if df.empty:
        raise ValueError(f"No data rows found in tab '{tab_name}'")

    # Normalize numeric fields where possible.
    for c in df.columns:
        # Coerce when a column is mostly numeric while preserving textual descriptor columns.
        try:
            converted = pd.to_numeric(df[c], errors="coerce")
            if converted.notna().mean() >= 0.8:
                df[c] = converted
        except Exception:
            continue

    return df


@dataclass
class RenderContext:
    wb: Workbook


# -------------------------
# Render mode implementations
# -------------------------

def _render_table(df: pd.DataFrame, title: str, output: str, options: Optional[Dict[str, Any]] = None) -> Path:
    fig, ax = plt.subplots(figsize=(16, 8))
    ax.axis("off")
    ax.set_title(title, fontsize=22, fontweight="bold", pad=18)
    table = ax.table(
        cellText=df.values.tolist(),
        colLabels=df.columns.tolist(),
        cellLoc="left",
        colLoc="left",
        loc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(11)
    table.scale(1.08, 2.0)

    for (r, c), cell in table.get_celld().items():
        cell.set_edgecolor("#d9d9d9")
        if r == 0:
            cell.set_facecolor("#0b3a63")
            cell.set_text_props(color="white", fontweight="bold")
        else:
            cell.set_facecolor("#f7f9fc" if r % 2 == 1 else "#ffffff")
            if c == 0:
                cell.set_text_props(fontweight="bold")

    return save_figure(fig, output)


def _render_timeline(df: pd.DataFrame, title: str, output: str, options: Optional[Dict[str, Any]] = None) -> Path:
    # Expect year in first col. Optional event labels from options.
    years = df.iloc[:, 0].astype(int).tolist()
    labels = (options or {}).get(
        "event_labels",
        {
            2006: "EU bans\nAGPs",
            2017: "US FDA VFD\nimplemented",
            2022: "EU bans\nzinc oxide",
            2024: "China AGP\ntightening",
            2025: "UK feed\nreform",
        },
    )

    fig, ax = plt.subplots(figsize=(11.8, 5.3))
    ax.axis("off")
    x_min, x_max = min(years) - 1, max(years) + 1
    ax.set_xlim(x_min, x_max)
    ax.set_ylim(-1.4, 1.4)

    ax.annotate("", xy=(x_max, 0), xytext=(x_min, 0), arrowprops=dict(arrowstyle="->", lw=2.0, color="#2e6ea3"))

    for i, y in enumerate(years):
        up = i % 2 == 0
        ax.vlines(y, 0, 0.45 if up else -0.45, color="#2e6ea3", linewidth=1.6)
        ax.scatter([y], [0], s=18, color="#0b3a63", zorder=3)
        txt = labels.get(y, str(y))
        ax.text(y, 0.55 if up else -0.6, txt, ha="center", va="center", fontsize=8)

    ax.set_title(title, fontsize=18, fontweight="bold", pad=10)
    return save_figure(fig, output)


def _render_scatter_cost_claim(df: pd.DataFrame, title: str, output: str, options: Optional[Dict[str, Any]] = None) -> Path:
    x = pd.to_numeric(df.iloc[:, 1], errors="coerce")
    y = pd.to_numeric(df.iloc[:, 3] if df.shape[1] > 3 else df.iloc[:, 2], errors="coerce")
    size_base = pd.to_numeric(df.iloc[:, 2], errors="coerce") if df.shape[1] > 3 else 1
    labels = df.iloc[:, 0].astype(str)

    fig, ax = plt.subplots(figsize=(10.5, 6.3))
    ax.scatter(x, y, s=np.clip(size_base, 0.2, None) * 40, color="#2e6ea3", alpha=0.9, edgecolor="white", linewidth=0.8)
    for xi, yi, label in zip(x, y, labels):
        ax.annotate(label, (xi, yi), textcoords="offset points", xytext=(4, 4), fontsize=8)

    ax.set_xscale("log")
    ax.set_xlabel(df.columns[1])
    ax.set_ylabel(df.columns[3] if df.shape[1] > 3 else df.columns[2])
    ax.set_title(title, fontsize=16, fontweight="bold")
    ax.grid(linestyle=":", alpha=0.3)
    return save_figure(fig, output)


def _render_innovation_matrix(df: pd.DataFrame, title: str, output: str, options: Optional[Dict[str, Any]] = None) -> Path:
    colors = {
        "Pharma-Nutra": "#0b3a63",
        "Ingredient Tech": "#1f77b4",
        "Brand/Consumer": "#2ca02c",
        "Feed/Health": "#ff7f0e",
        "Commodity Feed": "#d62728",
    }

    fig, ax = plt.subplots(figsize=(13.5, 8.8))
    for sector, part in df.groupby(df.columns[3]):
        ax.scatter(
            part.iloc[:, 1],
            part.iloc[:, 2],
            s=90,
            alpha=0.9,
            label=sector,
            color=colors.get(str(sector), "#7f7f7f"),
            edgecolor="white",
            linewidth=0.8,
        )

    for _, row in df.iterrows():
        ax.annotate(str(row.iloc[0]), (row.iloc[1], row.iloc[2]), textcoords="offset points", xytext=(4, 4), fontsize=8.6)

    ax.axvline(5, color="#4c4c4c", linestyle="--", linewidth=1)
    ax.axhline(20, color="#4c4c4c", linestyle="--", linewidth=1)
    ax.text(9.6, 39, "High R&D / High Margin", fontsize=9, color="#0b3a63", fontweight="bold")
    ax.text(1.0, 4.5, "Low R&D / Low Margin", fontsize=9, color="#6b6b6b")

    ax.set_xlabel(df.columns[1])
    ax.set_ylabel(df.columns[2])
    ax.set_title(title, fontsize=18, fontweight="bold")
    ax.grid(axis="both", linestyle=":", alpha=0.35)
    ax.legend(title=str(df.columns[3]), fontsize=8.5, title_fontsize=9.5, frameon=True, loc="lower right")
    return save_figure(fig, output)


def _render_grouped_bar(df: pd.DataFrame, title: str, output: str, options: Optional[Dict[str, Any]] = None) -> Path:
    opts = options or {}
    labels = df.iloc[:, 0].astype(str)
    cols = df.columns[1:]
    x = np.arange(len(labels))
    width = 0.36

    fig, ax = plt.subplots(figsize=(11.5, 6.5))
    for i, c in enumerate(cols):
        ax.bar(x + (i - (len(cols) - 1) / 2) * width, df[c].astype(float), width=width, label=c)

    rotation = float(opts.get("xtick_rotation", 12))
    ax.set_xticks(x, labels=labels)
    for tick in ax.get_xticklabels():
        tick.set_rotation(rotation)
        tick.set_ha("right" if rotation else "center")

    ax.set_xlabel(str(opts.get("xlabel", df.columns[0])))
    ax.set_ylabel(str(opts.get("ylabel", "Value")))
    ax.set_title(title, fontsize=18, fontweight="bold")
    if "ymin" in opts or "ymax" in opts:
        ymin = float(opts.get("ymin", ax.get_ylim()[0]))
        ymax = float(opts.get("ymax", ax.get_ylim()[1]))
        ax.set_ylim(ymin, ymax)
    ax.grid(axis="y", linestyle=":", alpha=0.35)
    ax.legend(loc=str(opts.get("legend_loc", "upper right")))
    return save_figure(fig, output)


def _render_dual_axis_bar_scatter(df: pd.DataFrame, title: str, output: str, options: Optional[Dict[str, Any]] = None) -> Path:
    cats = df.iloc[:, 0].astype(str)
    left = pd.to_numeric(df.iloc[:, 1], errors="coerce")
    right = pd.to_numeric(df.iloc[:, 2], errors="coerce")

    fig, ax1 = plt.subplots(figsize=(10.4, 6.2))
    ax1.bar(cats, left, color="#0b3a63", alpha=0.88, label=df.columns[1])
    ax1.set_ylabel(df.columns[1], color="#0b3a63")
    ax1.tick_params(axis="y", labelcolor="#0b3a63")
    ax1.tick_params(axis="x", rotation=35)
    ax1.grid(axis="y", linestyle=":", alpha=0.35)

    ax2 = ax1.twinx()
    ax2.scatter(cats, right, color="#d04a02", s=120, edgecolor="white", linewidth=1.2, label=df.columns[2], zorder=5)
    ax2.set_ylabel(df.columns[2], color="#d04a02")
    ax2.tick_params(axis="y", labelcolor="#d04a02")

    ax1.set_xlabel(str(df.columns[0]))
    ax1.set_title(title, fontsize=16, fontweight="bold")
    l1, lb1 = ax1.get_legend_handles_labels()
    l2, lb2 = ax2.get_legend_handles_labels()
    ax1.legend(l1 + l2, lb1 + lb2, loc="upper left", frameon=False)
    return save_figure(fig, output)


def _render_simple_bar(df: pd.DataFrame, title: str, output: str, horizontal: bool = False, options: Optional[Dict[str, Any]] = None) -> Path:
    opts = options or {}
    labels = df.iloc[:, 0].astype(str)
    values = pd.to_numeric(df.iloc[:, 1], errors="coerce")
    fig, ax = plt.subplots(figsize=tuple(opts.get("figsize", (10.2, 6.0))))
    bar_color = str(opts.get("bar_color", "#0b3a63"))
    if horizontal:
        ax.barh(labels, values, color=bar_color)
        ax.invert_yaxis()
        ax.set_xlabel(str(opts.get("xlabel", df.columns[1])))
        ax.set_ylabel(str(opts.get("ylabel", df.columns[0])))
        if opts.get("show_values"):
            dx = float(opts.get("value_dx", max(values) * 0.01 if len(values) else 0.1))
            suffix = str(opts.get("value_suffix", ""))
            for i, v in enumerate(values):
                ax.text(float(v) + dx, i, f"{float(v):.1f}{suffix}", va="center", fontsize=float(opts.get("value_fontsize", 9)), fontweight="bold", color=str(opts.get("value_color", "#2b2b2b")))
    else:
        ax.bar(labels, values, color=bar_color)
        ax.tick_params(axis="x", rotation=20)
        ax.set_xlabel(str(opts.get("xlabel", df.columns[0])))
        ax.set_ylabel(str(opts.get("ylabel", df.columns[1])))
    ax.set_title(title, fontsize=float(opts.get("title_size", 16)), fontweight="bold")
    ax.grid(axis="y" if not horizontal else "x", linestyle=":", alpha=0.35)
    return save_figure(fig, output)


def _render_segmented_barh(df: pd.DataFrame, title: str, output: str, options: Optional[Dict[str, Any]] = None) -> Path:
    opts = options or {}
    labels = df.iloc[:, 0].astype(str)
    values = pd.to_numeric(df.iloc[:, 1], errors="coerce")
    segments = df.iloc[:, 2].astype(str) if df.shape[1] > 2 else pd.Series(["All"] * len(df))

    palette = {
        "Pet Food & CPG": "#e67e22",
        "Pharma & Animal Health": "#2f455c",
        "Feed & Specialty Inputs": "#2daa5f",
    }

    fig, ax = plt.subplots(figsize=tuple(opts.get("figsize", (11.4, 7.1))))
    y = np.arange(len(labels))
    colors = [palette.get(s, "#4c6f86") for s in segments]
    ax.barh(y, values, color=colors)
    ax.set_yticks(y, labels=labels)
    ax.invert_yaxis()
    ax.set_xlabel(str(opts.get("xlabel", df.columns[1])))
    ax.set_ylabel(str(opts.get("ylabel", df.columns[0])))
    ax.set_title(title, fontsize=float(opts.get("title_size", 16)), fontweight="bold")
    ax.grid(axis="x", linestyle=":", alpha=0.35)

    xpad = float(np.nanmax(values) * 0.02) if len(values) else 0.2
    for yi, v in zip(y, values):
        ax.text(float(v) + xpad, yi, f"${float(v):.1f}B", va="center", fontsize=10, fontweight="bold")

    # Legend in deterministic order.
    from matplotlib.patches import Patch

    legend_order = ["Pet Food & CPG", "Pharma & Animal Health", "Feed & Specialty Inputs"]
    handles = [Patch(facecolor=palette[s], label=s) for s in legend_order if s in set(segments)]
    if handles:
        ax.legend(handles=handles, loc=str(opts.get("legend_loc", "lower right")), frameon=True)

    return save_figure(fig, output)


def _render_line(df: pd.DataFrame, title: str, output: str, options: Optional[Dict[str, Any]] = None) -> Path:
    opts = options or {}
    x = df.iloc[:, 0]
    fig, ax = plt.subplots(figsize=(10.2, 6.0))
    series = list(df.columns[1:])
    colors = opts.get("colors") or []
    linestyles = opts.get("linestyles") or []
    markers = opts.get("markers")
    linewidth = float(opts.get("linewidth", 2.2))
    for i, c in enumerate(series):
        y = pd.to_numeric(df[c], errors="coerce")
        kw: Dict[str, Any] = {
            "linewidth": linewidth,
            "label": str(opts.get("legend_labels", {}).get(c, c)),
            "linestyle": linestyles[i] if i < len(linestyles) else "-",
            "color": colors[i] if i < len(colors) else None,
        }
        if markers is not False:
            kw["marker"] = (markers[i] if isinstance(markers, list) and i < len(markers) else str(opts.get("marker", "o")))
            kw["markersize"] = float(opts.get("markersize", 3.5))
            kw["markeredgewidth"] = float(opts.get("markeredgewidth", 0))
        ax.plot(x, y, **kw)
    if opts.get("ymin") is not None or opts.get("ymax") is not None:
        ax.set_ylim(opts.get("ymin"), opts.get("ymax"))
    ax.set_xlabel(str(df.columns[0]))
    ax.set_ylabel(str(df.columns[1]) if len(df.columns) == 2 else "Value")
    if opts.get("ylabel"):
        ax.set_ylabel(str(opts["ylabel"]))
    if opts.get("xlabel"):
        ax.set_xlabel(str(opts["xlabel"]))
    ax.set_title(title, fontsize=float(opts.get("title_size", 16)), fontweight="bold")
    ax.grid(axis="y", linestyle=":", alpha=0.35)
    if len(df.columns) > 2:
        ax.legend(frameon=False, loc=str(opts.get("legend_loc", "best")))
    if opts.get("end_labels"):
        x_vals = pd.to_numeric(df.iloc[:, 0], errors="coerce")
        x_last = x_vals.iloc[-1] if not x_vals.isna().all() else len(x) - 1
        for i, c in enumerate(series):
            y = pd.to_numeric(df[c], errors="coerce")
            y_last = float(y.iloc[-1])
            label_fmt = str(opts.get("end_label_fmt", "{name} ({change:+.1f}%)"))
            change = float(y.iloc[-1] - y.iloc[0])
            txt = label_fmt.format(name=str(opts.get("legend_labels", {}).get(c, c)), change=change, value=y_last)
            ax.text(x_last + float(opts.get("end_label_dx", 0.10)), y_last + float(opts.get("end_label_dy", 0)), txt, fontsize=float(opts.get("end_label_size", 11)), fontweight="bold", color=(colors[i] if i < len(colors) else "#1f2937"), va="center")
    if opts.get("annotation"):
        ann = opts["annotation"]
        ax.annotate(
            ann.get("text", ""),
            xy=(ann.get("x"), ann.get("y")),
            xytext=(ann.get("text_x"), ann.get("text_y")),
            textcoords="data",
            fontsize=float(ann.get("fontsize", 9)),
            color=str(ann.get("color", "#111827")),
            arrowprops=dict(arrowstyle="->", lw=0.8, color=str(ann.get("arrow_color", "#111827"))),
        )
    return save_figure(fig, output)


def _render_area(df: pd.DataFrame, title: str, output: str) -> Path:
    x = pd.to_numeric(df.iloc[:, 0], errors="coerce")
    y = pd.to_numeric(df.iloc[:, 1], errors="coerce")
    fig, ax = plt.subplots(figsize=(10.2, 6.0))
    ax.fill_between(x, y, color="#2e6ea3", alpha=0.55)
    ax.plot(x, y, color="#0b3a63", linewidth=2)
    ax.set_xlabel(str(df.columns[0]))
    ax.set_ylabel(str(df.columns[1]))
    ax.set_title(title, fontsize=16, fontweight="bold")
    ax.grid(axis="y", linestyle=":", alpha=0.35)
    return save_figure(fig, output)


def _render_stacked_area(df: pd.DataFrame, title: str, output: str, options: Optional[Dict[str, Any]] = None) -> Path:
    opts = options or {}
    x = pd.to_numeric(df.iloc[:, 0], errors="coerce")
    ys = [pd.to_numeric(df[c], errors="coerce") for c in df.columns[1:]]
    fig, ax = plt.subplots(figsize=tuple(opts.get("figsize", (10.6, 6.0))))
    colors = opts.get("colors")
    ax.stackplot(x, ys, labels=df.columns[1:], alpha=float(opts.get("alpha", 0.85)), colors=colors)
    ax.set_xlabel(str(df.columns[0]))
    ax.set_ylabel(str(opts.get("ylabel", "Value")))
    ax.set_title(title, fontsize=float(opts.get("title_size", 16)), fontweight="bold")
    ax.grid(axis="y", linestyle=":", alpha=0.35)
    ax.legend(loc=str(opts.get("legend_loc", "upper right")), frameon=False, fontsize=float(opts.get("legend_size", 8)))
    return save_figure(fig, output)


def _render_stacked_column(df: pd.DataFrame, title: str, output: str) -> Path:
    x = np.arange(len(df.iloc[:, 0]))
    labels = df.iloc[:, 0].astype(str)
    fig, ax = plt.subplots(figsize=(10.6, 6.0))
    bottom = np.zeros(len(labels))
    for c in df.columns[1:]:
        vals = pd.to_numeric(df[c], errors="coerce")
        ax.bar(x, vals, bottom=bottom, label=c)
        bottom += np.nan_to_num(vals)
    ax.set_xticks(x, labels=labels)
    ax.set_xlabel(str(df.columns[0]))
    ax.set_ylabel("Value")
    ax.set_title(title, fontsize=16, fontweight="bold")
    ax.legend(frameon=False, fontsize=8)
    ax.grid(axis="y", linestyle=":", alpha=0.35)
    return save_figure(fig, output)


def _render_donut(df: pd.DataFrame, title: str, output: str, options: Optional[Dict[str, Any]] = None) -> Path:
    vals = pd.to_numeric(df.iloc[:, 1], errors="coerce")
    labels = df.iloc[:, 0].astype(str)
    opts = options or {}
    fig, ax = plt.subplots(figsize=tuple(opts.get("figsize", (8.6, 6.0))))
    colors = opts.get("colors")
    width = float(opts.get("ring_width", 0.42))
    pctdistance = float(opts.get("pctdistance", 0.82))
    startangle = float(opts.get("startangle", 90))
    wedges, texts, autotexts = ax.pie(
        vals,
        labels=labels,
        colors=colors,
        autopct="%1.1f%%",
        pctdistance=pctdistance,
        startangle=startangle,
        wedgeprops=dict(width=width, edgecolor="white"),
    )
    for t in autotexts:
        t.set_fontsize(float(opts.get("pct_fontsize", 8)))
        t.set_color("white")
    for t in texts:
        t.set_fontsize(float(opts.get("label_fontsize", 9)))
        t.set_color(str(opts.get("label_color", "#1f2937")))
    if opts.get("center_total_text"):
        ax.text(
            0,
            0,
            str(opts["center_total_text"]),
            ha="center",
            va="center",
            fontsize=float(opts.get("center_fontsize", 14)),
            fontweight="bold",
            color=str(opts.get("center_color", "#0b3a63")),
        )
    ax.set_title(title, fontsize=16, fontweight="bold")
    ax.set_aspect("equal")
    fig.tight_layout()
    return save_figure(fig, output)


def _render_smile_curve(df: pd.DataFrame, title: str, output: str, options: Optional[Dict[str, Any]] = None) -> Path:
    opts = options or {}
    # Use stage margins; overlay smooth curve and bars.
    x_lbl = df.iloc[:, 0].astype(str).tolist()
    y = pd.to_numeric(df.iloc[:, 1], errors="coerce").astype(float).to_numpy()
    x = np.arange(len(x_lbl))
    coeff = np.polyfit(x, y, 2)
    xp = np.linspace(0, len(x_lbl) - 1, 200)
    yp = np.polyval(coeff, xp)

    fig, ax = plt.subplots(figsize=tuple(opts.get("figsize", (10.8, 6.0))))
    if opts.get("show_bars", True):
        ax.bar(x_lbl, y, color=str(opts.get("bar_color", "#9fb8cc")), alpha=float(opts.get("bar_alpha", 0.55)), label="Observed")
    ax.plot(np.interp(xp, x, x), yp, color=str(opts.get("curve_color", "#2e6ea3")), linewidth=float(opts.get("curve_width", 2.5)), label="Smile-curve fit")
    point_color = str(opts.get("point_color", "#1788c3"))
    ax.scatter(x, y, s=float(opts.get("point_size", 220)), color=point_color, edgecolor="white", linewidth=1.2, zorder=5)
    if opts.get("point_labels", True):
        for xi, yi in zip(x, y):
            ax.text(xi, yi + float(opts.get("point_dy", 1.5)), f"{yi:.0f}%", color=point_color, ha="center", va="bottom", fontsize=float(opts.get("point_label_size", 12)), fontweight="bold")
    ax.set_xlabel(str(df.columns[0]))
    ax.set_ylabel(str(df.columns[1]))
    ax.set_title(title, fontsize=float(opts.get("title_size", 16)), fontweight="bold", loc=str(opts.get("title_loc", "center")))
    ax.grid(axis="y", linestyle=":", alpha=0.35)
    if opts.get("category_subtitles"):
        subtitles = opts["category_subtitles"]
        for i, txt in enumerate(subtitles):
            ax.text(i, ax.get_ylim()[0] - (ax.get_ylim()[1] - ax.get_ylim()[0]) * 0.05, txt, ha="center", va="top", fontsize=10, color="#6b7280", style="italic")
    if opts.get("show_guides", True):
        for xi, yi in zip(x, y):
            ax.vlines(xi, 0, yi, colors="#7c93ab", linestyles=":", linewidth=1)
    if opts.get("legend"):
        ax.legend(frameon=False)
    return save_figure(fig, output)


def _render_value_waterfall(df: pd.DataFrame, title: str, output: str) -> Path:
    labels = df.iloc[:, 0].astype(str)
    # Wrap long category labels to keep x-axis legible in DOCX scaling.
    labels_wrapped = labels.str.replace("/", "/\n", regex=False).str.replace(" ", "\n", n=1, regex=False)
    a = pd.to_numeric(df.iloc[:, 1], errors="coerce")
    b = pd.to_numeric(df.iloc[:, 2], errors="coerce")
    x = np.arange(len(labels_wrapped))

    fig, ax = plt.subplots(figsize=(12.2, 6.6))
    ax.bar(x - 0.18, a, width=0.36, label=df.columns[1], color="#0b3a63")
    ax.bar(x + 0.18, b, width=0.36, label=df.columns[2], color="#d04a02")
    ax.set_xticks(x, labels=labels_wrapped, rotation=12)
    ax.set_xlabel(str(df.columns[0]))
    ax.set_ylabel("Relative value split (index)")
    ax.set_title(title, fontsize=16, fontweight="bold")
    ax.legend(frameon=False)
    ax.grid(axis="y", linestyle=":", alpha=0.35)
    fig.tight_layout()
    return save_figure(fig, output)


def _render_risk_reward(df: pd.DataFrame, title: str, output: str, options: Optional[Dict[str, Any]] = None) -> Path:
    opts = options or {}
    labels = df.iloc[:, 0].astype(str)
    x = pd.to_numeric(df.iloc[:, 1], errors="coerce")
    y = pd.to_numeric(df.iloc[:, 2], errors="coerce")
    fig, ax = plt.subplots(figsize=tuple(opts.get("figsize", (10.6, 6.0))))
    ax.scatter(x, y, s=float(opts.get("marker_size", 160)), color=str(opts.get("marker_color", "#2e6ea3")), edgecolor="white", linewidth=1)
    for xi, yi, lbl in zip(x, y, labels):
        ax.annotate(lbl, (xi, yi), textcoords="offset points", xytext=(4, 4), fontsize=8)
    ax.set_xlabel(df.columns[1])
    ax.set_ylabel(df.columns[2])
    ax.set_title(title, fontsize=16, fontweight="bold")
    if opts.get("quadrants"):
        qx = float(opts.get("qx", np.nanmean(x)))
        qy = float(opts.get("qy", np.nanmean(y)))
        ax.axvline(qx, color="#6b7280", linestyle="--", linewidth=1)
        ax.axhline(qy, color="#6b7280", linestyle="--", linewidth=1)
        qlabels = opts.get("quadrant_labels") or {}
        if qlabels:
            xmin, xmax = ax.get_xlim()
            ymin, ymax = ax.get_ylim()
            ax.text((qx + xmax) / 2, (qy + ymax) / 2, str(qlabels.get("Q1", "")), ha="center", va="center", fontsize=9, color="#334155")
            ax.text((qx + xmax) / 2, (ymin + qy) / 2, str(qlabels.get("Q2", "")), ha="center", va="center", fontsize=9, color="#334155")
            ax.text((xmin + qx) / 2, (ymin + qy) / 2, str(qlabels.get("Q3", "")), ha="center", va="center", fontsize=9, color="#334155")
            ax.text((xmin + qx) / 2, (qy + ymax) / 2, str(qlabels.get("Q4", "")), ha="center", va="center", fontsize=9, color="#334155")
    ax.grid(linestyle=":", alpha=0.35)
    return save_figure(fig, output)


def _render_funnel(df: pd.DataFrame, title: str, output: str) -> Path:
    labels = df.iloc[:, 0].astype(str)
    vals = pd.to_numeric(df.iloc[:, 1], errors="coerce")
    fig, ax = plt.subplots(figsize=(10.5, 6.2))
    y = np.arange(len(labels))
    ax.barh(y, vals, color="#0b3a63")
    ax.set_yticks(y, labels=labels)
    ax.invert_yaxis()
    ax.set_title(title, fontsize=18, fontweight="bold")
    ax.set_xlabel(df.columns[1])
    ax.set_ylabel(df.columns[0])
    ax.grid(axis="x", linestyle=":", alpha=0.35)
    for i, v in enumerate(vals):
        ax.text(float(v) + max(vals) * 0.01, i, f"{int(v)}", va="center", fontsize=10)
    return save_figure(fig, output)


def _render_capability_matrix(df: pd.DataFrame, title: str, output: str, options: Optional[Dict[str, Any]] = None) -> Path:
    opts = options or {}
    labels_y = df.iloc[:, 0].astype(str)
    labels_x = df.columns[1:]
    numeric_df = df.iloc[:, 1:].apply(pd.to_numeric, errors="coerce")
    mat = numeric_df.to_numpy(dtype=float)
    fig_width = max(12.5, 0.60 * len(labels_x))
    fig, ax = plt.subplots(figsize=tuple(opts.get("figsize", (fig_width, 6.6))))

    core_color = str(opts.get("core_color", "#2f455c"))
    emerging_color = str(opts.get("emerging_color", "#9aaeb2"))
    none_color = str(opts.get("none_color", "#edf2f7"))
    core_threshold = float(opts.get("core_threshold", 0.75))
    emerging_threshold = float(opts.get("emerging_threshold", 0.25))
    marker_size = float(opts.get("marker_size", 260))

    for i in range(mat.shape[0]):
        for j in range(mat.shape[1]):
            val = mat[i, j]
            if np.isnan(val):
                color = none_color
                marker = ""
                text_color = "#5f6b7a"
            elif val >= core_threshold:
                color = core_color
                marker = ""
                text_color = "white"
            elif val >= emerging_threshold:
                color = emerging_color
                marker = ""
                text_color = "#213547"
            else:
                color = none_color
                marker = ""
                text_color = "#4c5b70"
            ax.scatter(j, i, s=marker_size, color=color, edgecolor="#1f2937", linewidth=0.6)
            if marker:
                ax.text(j, i, marker, ha="center", va="center", color=text_color, fontsize=8.5, fontweight="bold")

    ax.set_xticks(range(len(labels_x)), labels=labels_x, rotation=float(opts.get("x_rotation", 40)), ha="right")
    ax.tick_params(axis="x", labelsize=9.5)
    ax.set_yticks(range(len(labels_y)), labels=labels_y)
    ax.set_xlabel(str(opts.get("xlabel", "Company")))
    ax.set_ylabel(str(opts.get("ylabel", df.columns[0])))
    ax.set_title(title, fontsize=float(opts.get("title_size", 16)), fontweight="bold")
    ax.invert_yaxis()
    ax.grid(axis="both", linestyle="-", alpha=0.15)
    if opts.get("legend", True):
        from matplotlib.lines import Line2D

        handles = [
            Line2D([0], [0], marker="o", color="w", markerfacecolor=core_color, markeredgecolor="#1f2937", markersize=9, label="Core Competency"),
            Line2D([0], [0], marker="o", color="w", markerfacecolor=emerging_color, markeredgecolor="#1f2937", markersize=9, label="Emerging / Niche"),
        ]
        ax.legend(handles=handles, title="Status", loc=str(opts.get("legend_loc", "upper left")), bbox_to_anchor=tuple(opts.get("legend_anchor", (1.01, 1.0))), frameon=True)
    fig.tight_layout()
    return save_figure(fig, output)


def _render_multi_donut(df: pd.DataFrame, title: str, output: str, options: Optional[Dict[str, Any]] = None) -> Path:
    opts = options or {}
    # Expected format: first column = category/format, next columns = each species/value set.
    labels = df.iloc[:, 0].astype(str).tolist()
    groups = list(df.columns[1:])
    colors = opts.get("colors") or ["#0b3a63", "#1788c3", "#d04a02"]

    fig, axes = plt.subplots(1, len(groups), figsize=tuple(opts.get("figsize", (12.8, 4.3))))
    if len(groups) == 1:
        axes = [axes]

    for i, g in enumerate(groups):
        ax = axes[i]
        vals = pd.to_numeric(df[g], errors="coerce").fillna(0)
        wedges, texts, autotexts = ax.pie(
            vals,
            colors=colors[: len(vals)],
            startangle=140,
            autopct="%1.0f%%",
            pctdistance=0.78,
            wedgeprops=dict(width=0.42, edgecolor="white"),
            textprops=dict(color="#111827", fontsize=8),
        )
        for t in autotexts:
            t.set_color("white")
            t.set_fontsize(9)
            t.set_fontweight("bold")
        ax.set_title(str(opts.get("group_titles", {}).get(g, g)), fontsize=12, fontweight="bold", pad=8)
        ax.set_aspect("equal")

    fig.suptitle(title, fontsize=float(opts.get("title_size", 14)), fontweight="bold", y=0.98)
    fig.legend(labels, loc="lower center", ncol=min(3, len(labels)), frameon=False, bbox_to_anchor=(0.5, -0.02), fontsize=10)
    return save_figure(fig, output)


def _render_concentric(df: pd.DataFrame, title: str, output: str) -> Path:
    labels = df.iloc[:, 0].astype(str).tolist()
    vals = pd.to_numeric(df.iloc[:, 1], errors="coerce").tolist()
    fig, ax = plt.subplots(figsize=(9.8, 6.8))
    radii = [1.0, 0.6, 0.28]
    colors = ["#0b3a63", "#2e6ea3", "#d04a02"]

    for i, r in enumerate(radii[: len(vals)]):
        circle = plt.Circle((0, 0), r, color=colors[i], ec="white", lw=2)
        ax.add_artist(circle)

    label_x = 1.45
    for i, r in enumerate(radii[: len(vals)]):
        y = 0.72 - i * 0.5
        ax.plot([r * 0.72, label_x - 0.08], [y, y], color=colors[i], lw=1.6)
        ax.text(
            label_x,
            y,
            f"{labels[i]}: {vals[i]:.2f} bn",
            color="#1f2937",
            ha="left",
            va="center",
            fontsize=9.5,
            fontweight="bold",
        )

    ax.set_xlim(-1.2, 2.25)
    ax.set_ylim(-1.2, 1.2)
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title(title, fontsize=15, fontweight="bold")
    return save_figure(fig, output)


def _render_global_landscape(df: pd.DataFrame, title: str, output: str) -> Path:
    # Expected columns: Region, Category, Entity
    required = {"Region", "Category", "Entity"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Figure 46 tab missing columns: {', '.join(sorted(missing))}")

    map_pdf = ROOT / "sources" / "internal" / "MapChart_Map .pdf"
    if not map_pdf.exists():
        raise FileNotFoundError(f"Figure 46 background map is missing: {map_pdf}")

    try:
        import fitz  # PyMuPDF
    except Exception as exc:
        raise RuntimeError("PyMuPDF (fitz) is required to render Figure 46 from in-repo map PDF.") from exc

    doc = fitz.open(map_pdf)
    try:
        page = doc[0]
        images = page.get_images(full=True)
        if images:
            xref = images[0][0]
            extracted = doc.extract_image(xref)
            base = Image.open(io.BytesIO(extracted["image"])).convert("RGBA")
        else:
            pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0), alpha=False)
            base = Image.frombytes("RGB", [pix.width, pix.height], pix.samples).convert("RGBA")
    finally:
        doc.close()

    draw = ImageDraw.Draw(base, "RGBA")
    w, h = base.size

    # Typography tuned to preserve old-map readability across page insertions.
    fs_body = max(42, int(w * 0.0085))
    fs_cat = max(48, int(fs_body * 1.15))
    fs_region = max(58, int(fs_body * 1.28))

    def _font(candidates: List[str], size: int) -> Any:
        for c in candidates:
            try:
                return ImageFont.truetype(c, size=size)
            except Exception:
                continue
        return ImageFont.load_default()

    font_body = _font(
        [
            "/System/Library/Fonts/Helvetica.ttc",
            "/System/Library/Fonts/Supplemental/Arial.ttf",
        ],
        fs_body,
    )
    font_cat = _font(
        [
            "/System/Library/Fonts/Helvetica-Bold.ttc",
            "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
        ],
        fs_cat,
    )
    font_region = _font(
        [
            "/System/Library/Fonts/Helvetica-Bold.ttc",
            "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
        ],
        fs_region,
    )

    
    regions = ["North America", "UK", "Europe", "APAC", "LATAM"]
    cat_order = ["Corporates", "Startups", "Investors"]
    # Layout calibrated from legacy v19 map placement to preserve readability in-doc.
    anchors = {
        "North America": (0.005, 0.135),
        "UK": (0.335, 0.145),
        "Europe": (0.49, 0.13),
        "APAC": (0.79, 0.20),
        "LATAM": (0.30, 0.64),
    }
    widths = {
        "North America": 0.215,
        "UK": 0.15,
        "Europe": 0.205,
        "APAC": 0.195,
        "LATAM": 0.15,
    }

    for region in regions:
        subset_region = df[df["Region"].str.lower() == region.lower()]
        if subset_region.empty:
            continue

        x = int(anchors[region][0] * w)
        y = int(anchors[region][1] * h)
        bw = int(widths[region] * w)
        margin = max(8, int(w * 0.0045))
        line_gap = max(6, int(fs_body * 0.25))

        lines: List[tuple[str, str, ImageFont.ImageFont]] = []
        lines.append((region.upper(), "region", font_region))
        for cat in cat_order:
            rows = subset_region[subset_region["Category"].str.lower() == cat.lower()]
            if rows.empty:
                continue
            lines.append((cat.upper(), "cat", font_cat))
            for ent in rows["Entity"].astype(str).tolist():
                lines.append((f"• {ent}", "body", font_body))

        text_heights: List[int] = []
        for txt, kind, fnt in lines:
            bb = draw.textbbox((0, 0), txt, font=fnt)
            lh = bb[3] - bb[1]
            if kind in {"region", "cat"}:
                lh += line_gap * 2
            else:
                lh += line_gap
            text_heights.append(lh)

        bh = sum(text_heights) + margin * 2
        rect = (x, y, x + bw, y + bh)
        draw.rectangle(rect, fill=(255, 255, 255, 222), outline=(184, 195, 214, 255), width=max(1, int(w * 0.00035)))

        cy = y + margin
        for (txt, kind, fnt), lh in zip(lines, text_heights):
            if kind == "region":
                color = (31, 72, 154, 255)
                tx = x + margin
            elif kind == "cat":
                color = (157, 26, 24, 255)
                tx = x + margin
            else:
                color = (33, 33, 33, 255)
                tx = x + margin + int(fs_body * 0.25)
            draw.text((tx, cy), txt, font=fnt, fill=color)
            cy += lh

    out = EXPORTS_DIR / output
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    base = base.convert("RGB")
    base.save(out, format="PNG")
    _enforce_legacy_dimensions(out, output)
    return out


def _render_simplified_heatmap(df: pd.DataFrame, title: str, output: str) -> Path:
    # first col = species
    species = df.iloc[:, 0].astype(str).tolist()
    themes = df.columns[1:].tolist()
    matrix = df.iloc[:, 1:].apply(pd.to_numeric, errors="coerce").fillna(0).to_numpy(dtype=float)

    fig, ax = plt.subplots(figsize=(11.8, 6.5))
    im = ax.imshow(matrix, cmap="Blues", vmin=0, vmax=max(10, np.nanmax(matrix)))
    ax.set_xticks(np.arange(len(themes)), labels=themes)
    ax.set_yticks(np.arange(len(species)), labels=species)
    ax.set_title(title, fontsize=16, fontweight="bold")
    ax.set_xlabel(df.columns[0].replace("Primary ", "") if "Primary" in df.columns[0] else "Functional Theme")
    ax.set_ylabel(df.columns[0])

    for i in range(matrix.shape[0]):
        for j in range(matrix.shape[1]):
            val = int(matrix[i, j])
            ax.text(j, i, str(val), ha="center", va="center", color=("white" if val >= 7 else "#0b3a63"), fontsize=9)

    cbar = fig.colorbar(im, ax=ax, fraction=0.035, pad=0.02)
    cbar.set_label("Relative Strategic Relevance (0-10)")
    return save_figure(fig, output)


def _render_simplified_stacked(df: pd.DataFrame, title: str, output: str) -> Path:
    # Expected columns: Segment, Total, Companion/Pet-led, Production/Livestock-led
    cols = {c.lower(): c for c in df.columns}
    seg_col = next((c for c in df.columns if "segment" in c.lower()), df.columns[0])
    total_col = next((c for c in df.columns if "total" in c.lower()), df.columns[1])
    pet_col = next((c for c in df.columns if "pet" in c.lower() or "companion" in c.lower()), None)
    live_col = next((c for c in df.columns if "livestock" in c.lower() or "production" in c.lower()), None)
    if not pet_col or not live_col:
        raise ValueError("Simplified stacked tab must include pet/companion and livestock/production columns")

    seg = df[seg_col].astype(str)
    total = pd.to_numeric(df[total_col], errors="coerce")
    pet = pd.to_numeric(df[pet_col], errors="coerce")
    live = pd.to_numeric(df[live_col], errors="coerce")
    x = np.arange(len(seg))

    fig, ax = plt.subplots(figsize=(12.2, 6.8))
    ax.bar(x, pet, width=0.62, label=pet_col, color="#1f77b4")
    ax.bar(x, live, width=0.62, bottom=pet, label=live_col, color="#ff7f0e")

    for i, t in enumerate(total):
        if pd.notna(t):
            ax.text(i, float(t) + max(total) * 0.02, f"{int(t):,}", ha="center", va="bottom", fontsize=9)

    ax.set_xticks(x, labels=seg, rotation=15)
    ax.set_xlabel("Functional Segment")
    ax.set_ylabel("Estimated Value (USD millions)")
    ax.set_title(title, fontsize=16, fontweight="bold")
    ax.legend(loc="upper right")
    ax.grid(axis="y", linestyle=":", alpha=0.35)
    return save_figure(fig, output)


def dispatch_render(ctx: RenderContext, spec: Dict[str, Any]) -> Path:
    setup_style()
    tab = spec["tab"]
    mode = spec["mode"]
    output = spec["output_file"]
    title = spec.get("title") or tab
    options = spec.get("options") or {}

    df = read_df_from_tab(ctx.wb, tab)

    if mode == "table":
        return _render_table(df, title, output, options)
    if mode == "timeline":
        return _render_timeline(df, title, output, options)
    if mode == "scatter_cost_claim":
        return _render_scatter_cost_claim(df, title, output, options)
    if mode == "innovation_matrix":
        return _render_innovation_matrix(df, title, output, options)
    if mode == "grouped_bar":
        return _render_grouped_bar(df, title, output, options)
    if mode == "dual_axis_bar_scatter":
        return _render_dual_axis_bar_scatter(df, title, output, options)
    if mode == "bar_vertical":
        return _render_simple_bar(df, title, output, horizontal=False, options=options)
    if mode == "bar_horizontal":
        return _render_simple_bar(df, title, output, horizontal=True, options=options)
    if mode == "segmented_barh":
        return _render_segmented_barh(df, title, output, options)
    if mode == "line":
        return _render_line(df, title, output, options)
    if mode == "area":
        return _render_area(df, title, output)
    if mode == "stacked_area":
        return _render_stacked_area(df, title, output, options)
    if mode == "stacked_column":
        return _render_stacked_column(df, title, output)
    if mode == "donut":
        return _render_donut(df, title, output, options)
    if mode == "smile_curve":
        return _render_smile_curve(df, title, output, options)
    if mode == "value_waterfall":
        return _render_value_waterfall(df, title, output)
    if mode == "risk_reward":
        return _render_risk_reward(df, title, output, options)
    if mode == "funnel":
        return _render_funnel(df, title, output)
    if mode == "capability_matrix":
        return _render_capability_matrix(df, title, output, options)
    if mode == "multi_donut":
        return _render_multi_donut(df, title, output, options)
    if mode == "concentric":
        return _render_concentric(df, title, output)
    if mode == "global_landscape":
        return _render_global_landscape(df, title, output)
    if mode == "simplified_heatmap":
        return _render_simplified_heatmap(df, title, output)
    if mode == "simplified_stacked":
        return _render_simplified_stacked(df, title, output)

    raise ValueError(f"Unsupported render mode '{mode}' for {output}")
