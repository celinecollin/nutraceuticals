
from openpyxl import load_workbook
import json
import sys
import os

def read_registry(registry_path):
    """Reads the Sources tab of the registry and returns it as a list of dicts."""
    try:
        wb = load_workbook(registry_path, data_only=True)
        if 'Sources' not in wb.sheetnames:
             return {"error": "Sources tab not found"}
        
        sheet = wb['Sources']
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows:
            return []

        headers = rows[0]
        data = []
        for row in rows[1:]:
            # create dict from headers and row
            # Filter out None keys if headers has None
            item = {headers[i]: row[i] for i in range(len(headers)) if headers[i] is not None}
            # verify we have at least a source_id
            if item.get('source_id'):
                 data.append(item)
        
        return data
    except Exception as e:
        return {"error": str(e)}

if __name__ == "__main__":
    registry_path = "_registry/source_registry.xlsx"
    if len(sys.argv) > 1:
        registry_path = sys.argv[1]
    
    if not os.path.exists(registry_path):
        print(json.dumps({"error": f"File not found: {registry_path}"}))
        sys.exit(1)

    sources = read_registry(registry_path)
    print(json.dumps(sources, indent=2, default=str))
