import os
import subprocess
import openpyxl
from pathlib import Path

# Configuration
INPUT_DIR = "/Users/celinecollin/Library/CloudStorage/OneDrive-Personal/Nutraceuticals/report/20260121_new input"
OUTPUT_FILE = "/Users/celinecollin/Library/CloudStorage/OneDrive-Personal/Nutraceuticals/report/source_material/20260122_New_Input_Consolidated.md"

def extract_docx(file_path):
    """Extracts text from a .docx file using pandoc."""
    try:
        # Pandoc command to convert docx to markdown
        result = subprocess.run(
            ['pandoc', file_path, '-t', 'markdown'],
            capture_output=True,
            text=True,
            check=True
        )
        return result.stdout
    except subprocess.CalledProcessError as e:
        return f"Error extracting {file_path}: {e}\n{e.stderr}"

def extract_xlsx(file_path):
    """Extracts data from an .xlsx file using openpyxl and formats as markdown tables."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        markdown_output = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            markdown_output.append(f"### Sheet: {sheet_name}\n")
            
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                markdown_output.append("*Empty Sheet*\n")
                continue

            # Basic Markdown Table Construction
            # Using first row as header if available, otherwise just data
            
            # Filter out completely empty rows
            rows = [r for r in rows if any(cell is not None for cell in r)]
            
            if not rows:
                markdown_output.append("*Sheet contains no data*\n")
                continue

            header = rows[0]
            header_line = "| " + " | ".join(str(c) if c is not None else "" for c in header) + " |"
            separator_line = "| " + " | ".join("---" for _ in header) + " |"
            
            markdown_output.append(header_line)
            markdown_output.append(separator_line)
            
            for row in rows[1:]:
                row_line = "| " + " | ".join(str(c).replace('\n', ' ') if c is not None else "" for c in row) + " |"
                markdown_output.append(row_line)
            
            markdown_output.append("\n")
            
        return "\n".join(markdown_output)
    except Exception as e:
        return f"Error extracting {file_path}: {e}"

def main():
    print(f"Starting extraction from {INPUT_DIR}")
    files = sorted([f for f in os.listdir(INPUT_DIR) if not f.startswith('~')])
    
    consolidated_content = "# Consolidated New Input Materials (2026-01-21)\n\n"
    
    for filename in files:
        file_path = os.path.join(INPUT_DIR, filename)
        print(f"Processing {filename}...")
        
        consolidated_content += f"# Source: {filename}\n\n"
        
        if filename.endswith(".docx"):
            content = extract_docx(file_path)
            consolidated_content += content + "\n\n"
        elif filename.endswith(".xlsx"):
            content = extract_xlsx(file_path)
            consolidated_content += content + "\n\n"
        else:
            consolidated_content += f"*Skipped {filename} (unsupported format)*\n\n"
            
        consolidated_content += "---\n\n"
        
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write(consolidated_content)
        
    print(f"Consolidation complete. Output extracted to {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
